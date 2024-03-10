import openpyxl

def convertir_formulas_a_valores(archivo_excel):
    wb = openpyxl.load_workbook(archivo_excel)
    
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    formula = cell.value
                    if formula.startswith('='):
                        celda_referenciada = formula[1:]  # Eliminar el "=" del inicio de la fórmula
                        valor_referenciado = ws[celda_referenciada].value
                        cell.value = valor_referenciado

    wb.save(archivo_excel)
    print("Se han convertido todas las fórmulas a valores con éxito.")

archivo_excel = "electrohogar-venuts-gef.xlsx"  # Reemplaza "archivo.xlsx" con la ruta de tu archivo Excel
convertir_formulas_a_valores(archivo_excel)
